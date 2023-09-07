import os
import xlrd
import re
import pyperclip
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from rich import print


def init():
    print('检测更新中...')
    version = 1.9
    print('此程序版本：' + str(version))
    try:
        urlgithub = 'https://raw.githubusercontent.com/evilbutcher/Python/master/evanescent/release.json'
        update = requests.get(urlgithub, timeout=2)
        ver = update.json()
        if ver['releases'][0]['version'] > version:
            print('[bold yellow]更新[/bold yellow]啦！从GitHub获取更新详情成功！\n最新版本是：' +
                  str(ver['releases'][0]['version']))
            print('更新内容是：' + ver['releases'][0]['details'])
            print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
        else:
            print('检测更新完成，暂无更新\n')
    except (Exception):
        urlgitee = 'https://gitee.com/evilbutcher/Python/raw/master/evanescent/release.json'
        try:
            update = requests.get(urlgitee, timeout=2)
            ver = update.json()
            if ver['releases'][0]['version'] > version:
                print('[bold red]更新[/bold red]啦！从Gitee获取更新详情成功！\n最新版本是：' +
                      str(ver['releases'][0]['version']))
                print('更新内容是：' + ver['releases'][0]['details'])
                print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
            else:
                print('检测更新完成，暂无更新\n')
        except Exception as e:
            print('检测更新失败，原因：')
            print(str(e) + '\n')
    try:
        path = Path('origindata/')
        if path.is_dir() is False:
            os.mkdir(path)
        a = pyperclip.paste()
        if a == 'debugturnon':
            return True
    except Exception as e:
        print('初始化[bold red]失败[/bold red]，原因：' + str(e))


def dealxlsx(path: str, name: str, canprint: bool):
    try:
        mode = int(input('1 = 基线全为基准值\n2 = 基线仅在基准值处相同\n请输入“1或2”：'))
        if mode != 1 and mode != 2:
            print('模式输入错误，请重新运行程序重新输入！')
            return
        workbook = xlrd.open_workbook(path + '/' + name)
        sheet_name = workbook.sheet_names()[0]
        print('\n读取的Excel表名为：' + sheet_name)
        sheet = workbook.sheet_by_index(0)
        allrows = sheet.nrows
        allcols = sheet.ncols - 1
        print('一共有：' + str(allcols) + '列  ' + str(allrows) + '行（第一列时间未计入统计）')
        secend = input('请输入基准点（第几秒）：')
        valueforall = input('请输入一个基准值（自定即可）：')
        startpoint = int(secend)
        startvalue = sheet.cell(startpoint, 0).value
        print('所选取的修改起始点为：' + str(int(startvalue)) + '秒的数据')
        wb = Workbook()  # 引入result.xlsx工作表
        result = wb["Sheet"]
        if result.cell(1, 1).value is None:  # 因为两个库起始index不同，先设置一下第一行
            result.cell(1, 1).value = 0
            if canprint is True:
                print('\n1列1行 写入时间：0秒\n')
        for col in range(1, allcols + 1):  # 获取数据excel每列与基准值的差值
            minus = sheet.cell(startpoint, col).value - float(valueforall)
            print('第' + str(col) + '列截取的起始数据点为：' +
                  str(sheet.cell(startpoint, col).value))
            print('第' + str(col) + '列与基准值的差为：' + str(minus) + '\n')
            vcols = sheet.col_values(col)
            result.cell(1, col + 1).value = float(
                valueforall)  # 因为两个库起始index不同，先设置一下第一行
            if canprint is True:
                print(
                    str(col + 1) + '列1行' + '写入数据：' + str(float(valueforall)) +
                    '\n')
            for num in range(1, len(vcols)):  # 获取数据excel每行的数据，写入index+1的位置
                if result.cell(num + 1, 1).value is None:
                    result.cell(num + 1, 1).value = int(num)
                    if canprint is True:
                        print('1列' + str(num + 1) + '行 写入时间：' + str(int(num)) +
                              '秒' + '\n')
                if int(mode) == 1:
                    if num > startpoint:
                        if canprint is True:
                            print(
                                str(col + 1) + '列' + str(num + 1) + '行' +
                                '读取数据：' + str(sheet.cell(num, col).value))
                        if isinstance(sheet.cell(num, col).value,
                                      float) is True:
                            result.cell(
                                num + 1, col +
                                1).value = sheet.cell(num, col).value - minus
                        else:
                            print(
                                str(col) + '列' + str(num + 1) +
                                '行 读取数据[bold red]错误[/bold red]，请检查相应位置，此处选用上一个数值'
                            )
                            if isinstance(
                                    sheet.cell(num - 1, col).value,
                                    float) is True:
                                result.cell(num + 1,
                                            col + 1).value = sheet.cell(
                                                num - 1, col).value - minus
                            else:
                                print(
                                    str(col) + '列' + str(num + 1) +
                                    '行 读取数据出现[bold red]连续错误[/bold red]，请检查数据完整性'
                                )
                                return
                        if canprint is True:
                            print(
                                str(col + 1) + '列' + str(num + 1) + '行' +
                                '写入数据：' +
                                str(float(sheet.cell(num, col).value -
                                          minus)) + '\n')
                    else:
                        if canprint is True:
                            print(
                                str(col + 1) + '列' + str(num + 1) + '行' +
                                '读取数据：' + str(sheet.cell(num, col).value))
                        result.cell(num + 1,
                                    col + 1).value = float(valueforall)
                        if canprint is True:
                            print(
                                str(col + 1) + '列' + str(num + 1) + '行' +
                                '写入数据：' + str(float(valueforall)) + '\n')
                elif int(mode) == 2:
                    if canprint is True:
                        print(
                            str(col + 1) + '列' + str(num + 1) + '行' + '读取数据：' +
                            str(sheet.cell(num, col).value))
                    if isinstance(sheet.cell(num, col).value, float) is True:
                        result.cell(
                            num + 1,
                            col + 1).value = sheet.cell(num, col).value - minus
                    else:
                        print(
                            str(col) + '列' + str(num + 1) +
                            '行 读取数据[bold red]错误[/bold red]，请检查相应位置，此处选用上一个数值')
                        if isinstance(sheet.cell(num - 1, col).value,
                                      float) is True:
                            result.cell(num + 1, col + 1).value = sheet.cell(
                                num - 1, col).value - minus
                        else:
                            print(
                                str(col) + '列' + str(num + 1) +
                                '行 读取数据出现[bold red]连续错误[/bold red]，请检查数据完整性')
                            return
                    if canprint is True:
                        print(
                            str(col + 1) + '列' + str(num + 1) + '行' + '写入数据：' +
                            str(float(sheet.cell(num, col).value - minus)) +
                            '\n')
        wb.save("result.xlsx")
        print('[bold green]数据处理完成[/bold green]')
    except Exception as e:
        print('处理excel[bold red]失败[/bold red]，原因：' + str(e))


def setcolor(ws, num, name):
    rows = ws.max_row
    columns = ws.max_column
    color = 'FFB6C1'
    fille = PatternFill('solid', fgColor=color)
    waittocolor = []
    startnum = 1
    for m in range(0, num):
        startnum = startnum + 1
        waittocolor.append(startnum)
    for i in waittocolor:
        for k in range(i, columns + 1, num * 2):
            for j in range(1, rows + 1):
                ws.cell(j, k).fill = fille
    print(name + '[bold green]着色区分完成[/bold green]')
    return num


def setnumber():
    wb = load_workbook(r'result.xlsx')
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]
    num = int(input('请输入本次实验的平行实验次数：'))
    group = setcolor(ws, num, ws.title)  # 染色未排序表格
    wb.save(r'result.xlsx')
    print('如果无需排序请直接[bold red]关闭程序[/bold red]')
    goon = input('是否要继续排序？y=继续；n=不继续')
    if num == 1:
        print('只有一组数据无法进行组内排序，请退出程序')
    else:
        if goon == 'y':
            rows = ws.max_row
            columns = ws.max_column
            ws2 = wb.create_sheet('排序后数据', 0)
            for k in range(1, rows + 1):  # 生成时间列
                ws2.cell(k, 1).value = k - 1
            num = int(input('请输入要对比的第几秒的数据：'))
            for i in range(2, columns, group):  # 平行组内排序
                arry = []
                for j in range(i, i + group):  # 提取平行组内数据
                    arry.append(ws.cell(num + 1, j).value)
                arry.sort(reverse=True)
                print(arry)
                for n in range(i, i + group):  # 写入新表的相对位置
                    ws2.cell(num + 1, n).value = arry[n - i]
                for o in range(i, i + group):  # 开始比较
                    d = ws.cell(num + 1, o).value  # 表1中未排序的顺序值
                    for p in range(i, i + group):  # 在表2中循环比较匹配
                        d2 = ws2.cell(num + 1, p).value
                        if d2 == d:
                            for k in range(1, rows + 1):  # 全部行数
                                ws2.cell(k, p).value = ws.cell(k, o).value
            group = setcolor(ws2, group, ws2.title)  # 染色结果表格
        else:
            return
    wb.save(r'result.xlsx')


def main():
    try:
        canprint = False
        if init() is True:
            canprint = True
        name = input('请输入要处理的文件名')
        if re.search(r'xlsx|xls', name) is None:
            namewithsuffix4 = name + '.xlsx'
            path = Path('origindata/' + namewithsuffix4)
            if path.is_file() is True:
                dealxlsx('origindata/', namewithsuffix4, canprint)
            else:
                namewithsuffix3 = name + '.xls'
                path = Path('origindata/' + namewithsuffix3)
                if path.is_file() is True:
                    dealxlsx('origindata/', namewithsuffix3, canprint)
        else:
            dealxlsx('origindata/', name, canprint)
        exist = Path('result.xlsx')
        if exist.is_file() is True:
            setnumber()
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')
    except Exception as e:
        print('主函数运行[bold red]出现错误[/bold red]，原因：' + str(e))
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')


if __name__ == "__main__":
    main()
