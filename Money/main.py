from openpyxl import Workbook, load_workbook
from rich import print
from datetime import datetime
import pandas as pd


def main():
    try:
        wb = load_workbook("MOZE.xlsx")
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        rows = ws.max_row + 1
        for row in range(3, rows):
            ws.cell(row, 11).value = ws.cell(row,
                                             11).value.replace('/', '年', 1)
            ws.cell(row, 11).value = ws.cell(row, 11).value.replace('/', '月')
            ws.cell(row, 11).value = ws.cell(row, 11).value + "日 " + ws.cell(
                row, 12).value + ":00"
            date_string = ws.cell(row, 11).value
            date_object = datetime.strptime(date_string, "%Y年%m月%d日 %H:%M:%S")
            formatted_date_string = date_object.strftime("%Y年%m月%d日 %H:%M:%S")
            ws.cell(row, 12).value = formatted_date_string
        print("更改日期和时间成功")
        wb2 = Workbook()
        ws1 = wb2["Sheet"]  # 日期	类型	金额	一级分类	二级分类	账户1	账户2	备注   货币
        ws1.cell(1, 1).value = "日期"
        ws1.cell(1, 2).value = "类型"
        ws1.cell(1, 3).value = "金额"
        ws1.cell(1, 4).value = "一级分类"
        ws1.cell(1, 5).value = "二级分类"
        ws1.cell(1, 6).value = "账户1"
        ws1.cell(1, 7).value = "账户2"
        ws1.cell(1, 8).value = "备注"
        ws1.cell(1, 9).value = "货币"
        cols = ws1.max_column
        for row in range(3, rows):
            ws1.cell(row - 1, 1).value = ws.cell(row, 12).value  # 日期时间
            ws1.cell(row - 1, 2).value = ws.cell(row, 3).value  # 类型
            ws1.cell(row - 1, 3).value = abs(ws.cell(row, 6).value)  # 金额
            ws1.cell(row - 1, 4).value = ws.cell(row, 4).value  # 一级分类
            ws1.cell(row - 1, 5).value = ws.cell(row, 5).value  # 二级分类
            ws1.cell(row - 1, 6).value = ws.cell(row, 1).value  # 账户1
            ws1.cell(row - 1, 8).value = (str(ws.cell(row, 8).value) +
                                          str(ws.cell(row, 14).value)).replace(
                                              "None", "")  # 备注
            ws1.cell(row - 1, 9).value = ws.cell(row, 2).value  # 货币
        for row in range(3, rows):
            if ws1.cell(row - 1, 2).value == "转账":
                if ws1.cell(row - 1, 4).value == "转入" and ws1.cell(
                        row, 4).value == "转出":
                    ws1.cell(row - 1, 7).value = ws1.cell(row - 1, 6).value
                    ws1.cell(row - 1, 6).value = ws1.cell(row, 6).value
                    ws1.cell(row - 1, 4).value = "其他"
                elif ws1.cell(row - 1, 4).value == "转出" and ws1.cell(
                        row, 4).value == "转入":
                    ws1.cell(row - 1, 7).value = ws1.cell(row, 6).value
                    ws1.cell(row - 1, 4).value = "其他"
                elif (ws1.cell(row - 1, 4).value == "转出"
                      or ws1.cell(row - 1, 4).value == "转入") and ws1.cell(
                          row - 1, 3).value == 0:
                    for col in range(1, cols):
                        ws1.cell(row - 1, col).value = ""
                    print("第" + str(row - 1) + "行转账金额为零被清除")
                else:
                    print("第" + str(row - 1) + "行转账识别失败")
                for col in range(1, cols):
                    ws1.cell(row, col).value = ""
            if ws1.cell(row - 1, 2).value == "收入":
                ws1.cell(row - 1, 4).value = ws1.cell(row - 1, 5).value
                ws1.cell(row - 1, 5).value = ""
        print("icost创建信息成功")
        wb2.save("iCost.xlsx")
        # 读取Excel文件
        file_path = 'iCost.xlsx'
        df = pd.read_excel(file_path, engine='openpyxl')
        # 删除包含空值的行
        df = df.dropna(subset=['日期'])
        # 将更新后的DataFrame保存回Excel
        df.to_excel(file_path, index=False, engine='openpyxl')

    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


if __name__ == "__main__":
    main()
