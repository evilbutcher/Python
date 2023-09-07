from openpyxl import Workbook, load_workbook
from rich import print


def main():
    try:
        name = str(input("请输入文件名"))
        stop = int(input("请输入截止的核质比"))
        print("Loading...")
        wb = load_workbook(name + ".xlsx")
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        wb2 = Workbook()
        ws1 = wb2["Sheet"]
        rows = ws.max_row + 1
        for row in range(1, rows):
            if ws.cell(row, 1).value > stop:
                break
            if ws.cell(row, 2).value > 10000:
                ws1.cell(row, 2).value = ws.cell(row, 2).value
                ws1.cell(row, 1).value = ws.cell(row, 1).value
            if row % 3000 == 0:
                print("Processing：" + str(round(row / rows * 100, 2)) + "%")
        wb2.save("result.xlsx")
        print("Done!")
    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


if __name__ == "__main__":
    main()
