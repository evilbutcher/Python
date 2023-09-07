import os
import re
import pyperclip
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from rich import print


def init():
    print('检测更新中...')
    version = 1.4
    print('此程序版本：' + str(version))
    try:
        urlgithub = 'https://raw.githubusercontent.com/evilbutcher/Python/master/QPCR/release.json'
        update = requests.get(urlgithub, timeout=1)
        ver = update.json()
        if ver['releases'][0]['version'] > version:
            print('[bold yellow]更新[/bold yellow]啦！从GitHub获取更新详情成功！\n最新版本是：' +
                  str(ver['releases'][0]['version']))
            print('更新内容是：' + ver['releases'][0]['details'])
            print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
        else:
            print('检测更新完成，暂无更新\n')
    except (Exception):
        urlgitee = 'https://gitee.com/evilbutcher/Python/raw/master/QPCR/release.json'
        try:
            update = requests.get(urlgitee, timeout=1)
            ver = update.json()
            if ver['releases'][0]['version'] > version:
                print('[bold red]更新[/bold red]啦！从Gitee获取更新详情成功！\n最新版本是：' +
                      str(ver['releases'][0]['version']))
                print('更新内容是：' + ver['releases'][0]['details'])
                print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
            else:
                print('检测更新完成，暂无更新\n')
        except Exception as e:
            print('检测更新失败，原因：')
            print(str(e) + '\n')
    try:
        path = Path('origindata/')
        if path.is_dir() is False:
            os.mkdir(path)
        a = pyperclip.paste()
        if a == 'debugturnon':
            return True
    except Exception as e:
        print('初始化[bold red]失败[/bold red]，原因：' + str(e))


def dealxlsx(path: str, name: str, canprint: bool):
    try:
        wb = load_workbook(path + '/' + name)
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        rows = ws.max_row + 1
        slope = float(input('请输入Origin拟合方程的斜率：'))
        intercept = float(
            input('请输入Origin拟合方程的截距：'))  # Ct = intercept + slope * x
        ws.cell(1, 13).value = 'Log Copy'
        for row in range(2, rows):  # 计算Log Copy数
            try:
                samplename = str(ws.cell(row, 4).value)
                type = str(ws.cell(row, 5).value)
                num = float(ws.cell(row, 6).value)
                ws.cell(row, 13).value = (num - intercept) / slope
            except Exception as e:
                print('计算Log Copy出错，第' + str(row) + '行数据，名称：' + samplename +
                      '，类型：' + type + '，原因：' + str(e))
        ws.cell(1, 14).value = 'Copy'
        for row in range(2, rows):  # 计算Copy数
            try:
                samplename = str(ws.cell(row, 4).value)
                type = str(ws.cell(row, 5).value)
                index = float(ws.cell(row, 13).value)
                if samplename is not None:
                    if re.search(r'10$', samplename) is not None:
                        ws.cell(row, 14).value = 10**(index + 1)
                        if canprint is True:
                            print('第' + str(row) + '行正则匹配到10，则为稀释10倍，结果* 10')
                    elif re.search(r'100$', samplename) is not None:
                        ws.cell(row, 14).value = 10**(index + 2)
                        if canprint is True:
                            print('第' + str(row) +
                                  '行正则匹配到100，则为稀释100倍，结果* 100')
                    elif re.search(r'1000$', samplename) is not None:
                        ws.cell(row, 14).value = 10**(index + 3)
                        if canprint is True:
                            print('第' + str(row) +
                                  '行正则匹配到1000，则为稀释1000倍，结果* 1000')
                    elif re.search(r'10$|100$|1000$', samplename) is None:
                        ws.cell(row, 14).value = 10**index
                        if canprint is True:
                            print('第' + str(row) + '行名称不为空，正则未匹配到稀释倍数')
                else:
                    ws.cell(row, 14).value = 10**index
                    if canprint is True:
                        print('working curve')
                if canprint is True:
                    print(ws.cell(row, 14).value)
            except Exception as e:
                print('计算Copy出错，第' + str(row) + '行数据，名称：' + samplename +
                      '，类型：' + type + '，原因：' + str(e))
        WashVolumecoefficient = float(input('请输入清洗液的体积（单位：μL）：')) / 2
        ElutionVolumecoefficient = float(input('请输入洗脱液的体积（单位：μL）：')) / 2
        ws.cell(1, 15).value = 'DNA(pmole)'
        for row in range(2, rows):  # 计算mol数
            try:
                samplename = str(ws.cell(row, 4).value)
                type = str(ws.cell(row, 5).value)
                if samplename is not None and samplename != 'None':
                    Copy = float(ws.cell(row, 14).value)
                    if re.search(r'w|W|wash|Wash', samplename) is not None:
                        ws.cell(row, 15).value = Copy / (
                            6.02 * 10**23) * WashVolumecoefficient * 10**12
                        if canprint is True:
                            print('处理' + samplename +
                                  '[bold green]识别成功[/bold green]')
                    elif re.search(r'r|elu|elution|R',
                                   samplename) is not None:
                        ws.cell(row, 15).value = Copy / (
                            6.02 * 10**23) * ElutionVolumecoefficient * 10**12
                        if canprint is True:
                            print('处理' + samplename +
                                  '[bold green]识别成功[/bold green]')
                    else:
                        print(
                            '第' + str(row) +
                            '行样品名称未能识别，请手动转换，转换时注意[bold red]体积倍数[/bold red]关系！'
                        )
            except Exception as e:
                print('转换pmol出错，第' + str(row) + '行数据，名称：' + samplename +
                      '，类型：' + type + '，原因：' + str(e))
        ws.cell(1, 16).value = 'Averange DNA(pmole)'
        # 不再需要 ws.cell(1, 17).value = 'Final Averange DNA(pmole)'
        for row in range(2, rows, 3):  # 计算平均mol数
            try:
                start = str(ws.cell(row, 5).value)
                if start == "Unknown Sample":
                    pmol1 = float(ws.cell(row, 15).value)
                    pmol2 = float(ws.cell(row + 1, 15).value)
                    pmol3 = float(ws.cell(row + 2, 15).value)
                    ws.cell(row + 1, 16).value = (pmol1 + pmol2 + pmol3) / 3
                '''不再需要
                pmol4 = float(ws.cell(row + 3, 15).value)
                pmol5 = float(ws.cell(row + 4, 15).value)
                pmol6 = float(ws.cell(row + 5, 15).value)
                ws.cell(row + 4, 16).value = (pmol4 + pmol5 + pmol6) / 3
                ws.merge_cells(start_row=row + 2,
                               end_row=row + 3,
                               start_column=17,
                               end_column=17)
                ws.cell(row + 2, 17).value = (ws.cell(row + 1, 16).value +
                                              ws.cell(row + 4, 16).value) / 2
                '''
            except Exception as e:
                print('计算pmol平均值出错，第' + str(row) + '行数据，原因：' + str(e))
        for row in range(2, rows, 6):  # 着色区分
            columns = ws.max_column + 1
            color = 'FFB6C1'
            fille = PatternFill('solid', fgColor=color)
            for column in range(1, columns):
                ws.cell(row, column).fill = fille
                ws.cell(row + 1, column).fill = fille
                ws.cell(row + 2, column).fill = fille
        print('[bold green]着色区分完成[/bold green]')
        wb.save(path + name)
        print('[bold green]数据处理完成[/bold green]')
    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


def main():
    try:
        canprint = False
        if init() is True:
            canprint = True
        name = input('请输入要处理的文件名')
        if re.search(r'xlsx|xls', name) is None:
            namewithsuffix4 = name + '.xlsx'
            path = Path('origindata/' + namewithsuffix4)
            if path.is_file() is True:
                dealxlsx('origindata/', namewithsuffix4, canprint)
            else:
                namewithsuffix3 = name + '.xls'
                path = Path('origindata/' + namewithsuffix3)
                if path.is_file() is True:
                    dealxlsx('origindata/', namewithsuffix3, canprint)
        else:
            dealxlsx('origindata/', name, canprint)
            print('here')
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')
    except Exception as e:
        print('主函数运行[bold red]出现错误[/bold red]，原因：' + str(e))
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')


if __name__ == "__main__":
    main()
