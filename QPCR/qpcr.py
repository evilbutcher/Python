import os
import re
import pyperclip
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from rich import print


def init():
    print('检测更新中...')
    version = 10.9
    print('此程序版本：' + str(version))
    try:
        urlgithub = 'https://raw.githubusercontent.com/evilbutcher/Python/master/QPCR/release.json'
        update = requests.get(urlgithub, timeout=2)
        ver = update.json()
        if ver['releases'][0]['version'] > version:
            print('[bold yellow]更新[/bold yellow]啦！从GitHub获取更新详情成功！\n最新版本是：' +
                  str(ver['releases'][0]['version']))
            print('更新内容是：' + ver['releases'][0]['details'])
            print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
        else:
            print('检测更新完成，暂无更新\n')
    except (Exception):
        urlgitee = 'https://gitee.com/evilbutcher/Python/raw/master/QPCR/release.json'
        try:
            update = requests.get(urlgitee, timeout=2)
            ver = update.json()
            if ver['releases'][0]['version'] > version:
                print('[bold red]更新[/bold red]啦！从Gitee获取更新详情成功！\n最新版本是：' +
                      str(ver['releases'][0]['version']))
                print('更新内容是：' + ver['releases'][0]['details'])
                print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
            else:
                print('检测更新完成，暂无更新\n')
        except Exception as e:
            print('检测更新失败，原因：')
            print(str(e) + '\n')
    try:
        path = Path('origindata/')
        if path.is_dir() is False:
            os.mkdir(path)
        a = pyperclip.paste()
        if a == 'debugturnon':
            return True
    except Exception as e:
        print('初始化[bold red]失败[/bold red]，原因：' + str(e))


def dealxlsx(path: str, name: str, canprint: bool):
    try:
        wb = load_workbook(path + '/' + name)
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        rows = ws.max_row
        columns = ws.max_column
        slope = float(input('请输入Origin拟合方程的斜率'))
        intercept = float(input('请输入Origin拟合方程的截距'))
    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


def main():
    try:
        canprint = False
        if init() is True:
            canprint = True
        name = input('请输入要处理的文件名')
        if re.search(r'xlsx|xls', name) is None:
            namewithsuffix4 = name + '.xlsx'
            path = Path('origindata/' + namewithsuffix4)
            if path.is_file() is True:
                dealxlsx('origindata/', namewithsuffix4, canprint)
            else:
                namewithsuffix3 = name + '.xls'
                path = Path('origindata/' + namewithsuffix3)
                if path.is_file() is True:
                    dealxlsx('origindata/', namewithsuffix3, canprint)
        else:
            dealxlsx('origindata/', name, canprint)
            print('here')
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')
    except Exception as e:
        print('主函数运行[bold red]出现错误[/bold red]，原因：' + str(e))
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')


if __name__ == "__main__":
    main()
