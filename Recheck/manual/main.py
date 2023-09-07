from openpyxl import Workbook, load_workbook
from rich import print


def main():
    try:
        filepeople = str(input("请输入人员表文件名"))
        print("Loading...")
        wb = load_workbook(filepeople + ".xlsx")
        name = input('请输入阴性人员姓名，以中文逗号分隔：')
        print("[bold green]待更新人员姓名[/bold green] "+name)
        sheetnames = wb.sheetnames
        names = name.split('，')
        for i in range(0, len(sheetnames)):  # 循环历遍所有sheet表
            print("[bold red]当前识别[/bold red]  "+sheetnames[i])
            ws = wb[sheetnames[i]]
            rows = ws.max_row  # 获取人员表最大行
            for name in names:  # 循环对比人名
                for row in range(1, rows):
                    if ws.cell(row, 6).value == name:
                        ws.cell(row, 22).value = "阴性"
                        print(name + "[bold green]替换完成[/bold green]")
        wb.save(filepeople + "（已更新核酸结果）.xlsx")
        input('请按任意键退出')
    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


if __name__ == "__main__":
    main()
