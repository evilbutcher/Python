from openpyxl import Workbook, load_workbook
from rich import print
import pyperclip

def main():
    try:
        filepeople = str(input("请输入人员表文件名"))
        filename = str(input("请输入姓名表文件名"))
        print("Loading...")
        wb = load_workbook(filepeople + ".xlsx")
        wbname = load_workbook(filename + ".xlsx")
        sheet2 = wbname.sheetnames  # 读取姓名表所有sheet名
        wsname = wbname[sheet2[0]]  # 读取姓名表第一个sheet表
        rows2 = wsname.max_row  # 读取姓名表最大行
        print('正在获取姓名...')
        name = ""  # 设定名称字符串变量
        for i in range(2, rows2+1):
            if i == 2:
                name = name + wsname.cell(i, 7).value  # 以中文逗号拼接姓名
            else:
                name = name + "，"+wsname.cell(i, 7).value  # 以中文逗号拼接姓名
        print("[bold green]待更新人员姓名[/bold green] "+name)
        sheetnames = wb.sheetnames
        names = name.split('，')
        notfoundnames = []
        for name in names:  # 循环对比人名
            print("[bold red]当前检索姓名[/bold red]：" + name)
            flag = 0
            for i in range(0, len(sheetnames)):  # 循环历遍所有sheet表
                print("正在搜索 " + sheetnames[i])
                ws = wb[sheetnames[i]]
                rows = ws.max_row  # 获取人员表最大行
                for row in range(1, rows):
                    if ws.cell(row, 6).value == name:
                        ws.cell(row, 22).value = "阴性"
                        print(name + "[bold green]替换完成[/bold green]")
                        flag = 1
            if flag == 0:
                notfoundnames.append(name)
                print(name + "[bold yellow]检索失败[/bold yellow]")
        if len(notfoundnames) == 0:
            print('人员已全部更新')
        else:
            print("以下人员[bold red]未在表中检索到[/bold red]，请检查姓名拼写，已将下列姓名复制进系统剪贴板，[bold green]可在微信中直接粘贴[/bold green]")
            result = ""
            for notfoundname in notfoundnames:
                result = result + " " + notfoundname
            print(result)
            pyperclip.copy(result)
        wb.save(filepeople + "（已更新核酸结果）.xlsx")
        input('请按任意键退出')
    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


if __name__ == "__main__":
    main()
