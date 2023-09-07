from openpyxl import Workbook, load_workbook
from rich import print
import re


def main():
    try:
        print("R1载入中...")
        wb1 = load_workbook("R1.xlsx")
        R1 = wb1["Sheet1"]
        print("R1载入完成")
        print("R2载入中...")
        wb2 = load_workbook("R2.xlsx")
        R2 = wb2["Sheet1"]
        print("R2载入完成")
        print("R3载入中...")
        wb3 = load_workbook("R3.xlsx")
        R3 = wb3["Sheet1"]
        print("R3载入完成")
        print("R4载入中...")
        wb4 = load_workbook("R4.xlsx")
        R4 = wb4["Sheet1"]
        print("R4载入完成")
        r1rows = R1.max_row + 1
        r2rows = R2.max_row + 1
        r3rows = R3.max_row + 1
        # r4rows = R4.max_row + 1
        wb = Workbook()
        ws = wb["Sheet"]
        ws.cell(1, 1).value = "R4 Count"
        ws.cell(1, 2).value = "序列"
        ws.cell(1, 3).value = "R3"
        ws.cell(1, 4).value = "R3 Count"
        ws.cell(1, 5).value = "R2"
        ws.cell(1, 6).value = "R2 Count"
        ws.cell(1, 7).value = "R1"
        ws.cell(1, 8).value = "R1 Count"
        for row in range(2, 10000):
            if int(R4.cell(row, 1).value) > 2:
                candidate = R4.cell(row, 4).value
                ws.cell(row, 2).value = candidate
                ws.cell(row, 1).value = R4.cell(row, 1).value
                nin3 = 0
                for i in range(2, r3rows):
                    if re.search(str(candidate),
                                 str(R3.cell(i, 4).value)) is not None:
                        nin3 = nin3 + 1
                        ws.cell(row,
                                4).value = ws.cell(row, 4).value + R3.cell(
                                    i, 1).value
                ws.cell(row, 3).value = nin3
                print("R3-" + str(row) + "行统计完成")
                nin2 = 0
                for j in range(2, r2rows):
                    if re.search(str(candidate),
                                 str(R2.cell(j, 4).value)) is not None:
                        nin2 = nin2 + 1
                        ws.cell(row,
                                6).value = ws.cell(row, 6).value + R2.cell(
                                    i, 1).value
                ws.cell(row, 5).value = nin2
                print("R2-" + str(row) + "行统计完成")
                nin1 = 0
                for k in range(2, r1rows):
                    if re.search(str(candidate),
                                 str(R1.cell(k, 4).value)) is not None:
                        nin1 = nin1 + 1
                        ws.cell(row,
                                8).value = ws.cell(row, 8).value + R1.cell(
                                    i, 1).value
                ws.cell(row, 7).value = nin1
                print("R1-" + str(row) + "行统计完成")
        wb.save("Result.xlsx")
    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


if __name__ == "__main__":
    main()
