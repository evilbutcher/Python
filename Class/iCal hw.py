from openpyxl import Workbook, load_workbook
from rich import print
import re
from icalendar import Calendar, Event, Alarm
from hashlib import md5
from datetime import datetime, timedelta, date, time


def deal():
    try:
        numberofweek = int(input('请输入本课表持续的周数：'))
        wb = load_workbook("kb.xlsx")
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        mycalender = Calendar()  # 创建日历
        mycalender.add('X-WR-CALNAME', '课程表')  # 添加属性
        mycalender.add('prodid', '-//My calendar//luan//CN')
        mycalender.add('version', '2.0')
        mycalender.add('METHOD', 'PUBLISH')
        mycalender.add('CALSCALE', 'GREGORIAN')  # 历法：公历
        mycalender.add('X-WR-TIMEZONE', 'Asia/Shanghai')  # 通用扩展属性，表示时区
        for w in range(2, 7):
            week = w-1
            for i in range(3, 12):
                num = i-2
                if ws.cell(i, w).value != '' and ws.cell(i, w).value != None:
                    courses = ws.cell(i, w).value
                    if num > 5:
                        listnum = num - 1
                    else:
                        listnum = num
                    if courses.find('\n') != -1:
                        newcourses = courses.split('\n')
                        for course in newcourses:
                            coursename = course[2:]
                            print(coursename + ' 星期'+str(week) +
                                  ' 第'+str(listnum)+'节')
                            creatics(mycalender, numberofweek,
                                     coursename, week, listnum)
                    else:
                        coursename = courses[2:]
                        print(coursename + ' 星期'+str(week) +
                              ' 第'+str(listnum)+'节')
                        creatics(mycalender, numberofweek,
                                 coursename, week, listnum)
        with open('课程表.ics', 'wb') as file:
            file.write(mycalender.to_ical().replace(b'\r\n', b'\n').strip())
            print('创建完毕！')

    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


def uid_generate(k): return md5(k.encode("utf-8")).hexdigest()


def getday(startDate, number): return startDate + timedelta(days=int(number-1))


def creatics(mycalender, numberofweek, coursename, week, listnum):
    try:
        startDate = date(2023, 9, 4)  # 开学第一周星期一的日期，存储为年、月、日三项
        classTime = [time(8, 5, 0), time(9, 0, 0), time(10, 10, 0), time(11, 40, 0),
                     time(13, 00, 0), time(13, 55, 0), time(14, 45, 0), time(15, 35, 0)]
        startDay = getday(startDate, week)
        starttime = classTime[listnum-1]
        finalstarttime = datetime.combine(startDay, starttime)
        print(finalstarttime)
        finalendtime = finalstarttime + timedelta(minutes=40)  # 每一节课的时长分钟数

        event = Event()  # 添加事件
        event.add('uid', uid_generate(coursename + str(finalstarttime)))
        event.add('summary', coursename)
        event.add('dtstart', finalstarttime)
        event.add('dtend', finalendtime)
        event.add('dtstamp', datetime.now())
        event.add('rrule', {'freq': 'weekly', 'count': numberofweek})  # 重复周次

        alarm = Alarm()  # 编辑提醒
        alarm.add('action', 'DISPLAY')
        alarm.add('description', '要上课了！'+coursename)
        alarm.add('trigger;related=start', '-PT30M')
        event.add_component(alarm)

        mycalender.add_component(event)  # 将事件添加到日历中

    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


if __name__ == "__main__":
    deal()
