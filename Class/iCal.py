from openpyxl import Workbook, load_workbook
from rich import print
import re
from icalendar import Calendar, Event, Alarm
from hashlib import md5
from datetime import datetime, timedelta, date, time


def deal():
    try:
        numberofweek = int(input('请输入本课表持续的周数：'))
        wb = load_workbook("kb.xlsx")
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        rows = ws.max_row
        person = int(rows) / 11
        mycalender = Calendar()  # 创建日历
        mycalender.add('X-WR-CALNAME', '课程表')  # 添加属性
        mycalender.add('prodid', '-//My calendar//luan//CN')
        mycalender.add('version', '2.0')
        mycalender.add('METHOD', 'PUBLISH')
        mycalender.add('CALSCALE', 'GREGORIAN')  # 历法：公历
        mycalender.add('X-WR-TIMEZONE', 'Asia/Shanghai')  # 通用扩展属性，表示时区
        for i in range(0, int(person)):
            print('[bold green]正在获取姓名：[/bold green]')
            namerow = ws.cell(1+i*11, 1).value  # 获取姓名行
            getname = re.match('.*?【', namerow)  # 正则匹配姓名
            teachername = getname[0][:-1]  # 获取姓名
            print(teachername)
            for w in range(2, 7):
                for j in range(1+i*11, 11+i*11):
                    if ws.cell(j, w).value != '' and ws.cell(j, w).value != None and ws.cell(j, w).value.startswith('星') == False:
                        course = ws.cell(j, w).value
                        listnum = int((j % 11)-2)
                        week = int(w-1)
                        print(course+' '+'星期' + str(week) +
                              ' 第'+str(listnum)+'节课')
                        creatics(mycalender, numberofweek, course, week,
                                 listnum, teachername)
        with open('课程表.ics', 'wb') as file:
            file.write(mycalender.to_ical().replace(b'\r\n', b'\n').strip())
            print('创建完毕！')

    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


def uid_generate(k): return md5(k.encode("utf-8")).hexdigest()


def getday(startDate, number): return startDate + timedelta(days=int(number-1))


def creatics(mycalender, numberofweek, course, week, listnum, teachername):
    try:
        startDate = date(2023, 9, 4)  # 开学第一周星期一的日期，存储为年、月、日三项
        northclassTime = [time(8, 0, 0), time(8, 50, 0), time(9, 40, 0), time(10, 50, 0),
                          time(11, 40, 0), time(13, 30, 0), time(14, 20, 0), time(15, 10, 0)]  # 北每节课的上课时间
        southclassTime = [time(8, 0, 0), time(8, 50, 0), time(10, 0, 0), time(10, 50, 0),
                          time(11, 40, 0), time(13, 30, 0), time(14, 20, 0), time(15, 10, 0)]  # 南每节课的上课时间
        startDay = getday(startDate, week)
        if str(course).startswith('北') == True:
            starttime = northclassTime[listnum-1]
            finalstarttime = datetime.combine(startDay, starttime)
        else:
            starttime = southclassTime[listnum-1]
            finalstarttime = datetime.combine(startDay, starttime)
        print(finalstarttime)
        finalendtime = finalstarttime + timedelta(minutes=40)  # 每一节课的时长分钟数
        if course.find('研') != -1:
            location = course
        else:
            location = course[:-2]

        event = Event()  # 添加事件
        event.add('uid', uid_generate(teachername + str(finalstarttime)))
        event.add('summary', teachername)
        event.add('dtstart', finalstarttime)
        event.add('dtend', finalendtime)
        event.add('dtstamp', datetime.now())
        event.add('location', location)
        event.add('rrule', {'freq': 'weekly', 'count': numberofweek})  # 重复周次

        alarm = Alarm()  # 编辑提醒
        alarm.add('action', 'DISPLAY')
        alarm.add('description', '要上课了！'+location)
        alarm.add('trigger;related=start', '-PT30M')
        event.add_component(alarm)

        mycalender.add_component(event)  # 将事件添加到日历中

    except Exception as e:
        print('处理数据[bold red]出现错误[/bold red]，原因：' + str(e))


if __name__ == "__main__":
    deal()
